from __future__ import annotations

import tempfile
import unittest
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook

from marketing_diagnosis.customer_excel_result import enrich_customer_excel_result
from marketing_diagnosis.excel_loader import load_excel_dataset


class CustomerExcelLoaderTests(unittest.TestCase):
    def _save(self, workbook: Workbook) -> Path:
        handle = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        handle.close()
        path = Path(handle.name)
        workbook.save(path)
        self.addCleanup(path.unlink, missing_ok=True)
        return path

    def test_compact_chinese_template_reads_only_imported_rows(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "02_基础经营"
        sheet.append(["02_基础经营"])
        sheet.append(["说明"])
        sheet.append([])
        sheet.append(["A. 酒店基本信息"])
        sheet.append(["说明"])
        sheet.append(["是否导入", "系统模块（勿改）", "酒店ID", "酒店名称", "诊断开始日期", "诊断结束日期"])
        sheet.append(["是", "basic_info", "hotel_001", "测试酒店", date(2026, 6, 16), date(2026, 7, 15)])
        sheet.append([])
        sheet.append(["B. 日度经营数据"])
        sheet.append(["说明"])
        sheet.append(["是否导入", "系统模块（勿改）", "酒店ID", "酒店名称", "业务日期", "总房量", "售出间夜", "房费收入", "平均房价", "出租率", "快照时间"])
        sheet.append(["是", "hotel_daily", "hotel_001", "测试酒店", date(2026, 7, 15), 32, 28, 4060, 145, .875, datetime(2026, 7, 15, 23, 50)])
        sheet.append(["否", "hotel_daily", "demo_hotel", "不应导入", date(2026, 7, 14), 32, 1, 1, 1, .01, datetime(2026, 7, 14, 23, 50)])

        flow = workbook.create_sheet("04_流量用户")
        flow.append(["是否导入", "系统模块（勿改）", "酒店ID", "业务日期", "平台", "统计周期", "曝光人数", "同行曝光", "曝光排名", "浏览人数", "同行浏览", "浏览排名", "支付订单", "同行订单", "订单排名", "曝光-浏览转化率", "同行曝光-浏览转化率", "转化率排名", "浏览-支付转化率", "同行浏览-支付转化率", "支付转化排名", "HOS分", "HOS排名", "信息分", "快照时间"])
        flow.append(["是", "ota_funnel", "hotel_001", date(2026, 7, 15), "美团", "日", 2200, 2250, "3/20", 225, 205, "4/20", 18, 20, "5/20", .1023, .0911, "6/20", .08, .0976, "7/20", 5.35, "2/20", 95, datetime(2026, 7, 15, 23, 45)])

        performance = workbook.create_sheet("经营同比")
        performance.append(["是否导入", "系统模块（勿改）", "酒店ID", "酒店名称", "业务日期", "指标名称", "当日值", "本月值", "本年值", "快照时间"])
        performance.append(["是", "hotel_performance_daily", "hotel_001", "测试酒店", date(2026, 7, 15), "出租率", .92, .905, .863, datetime(2026, 7, 15, 18)])

        dataset = load_excel_dataset(self._save(workbook))

        self.assertEqual(dataset["__excel_context__"]["hotel_id"], "hotel_001")
        self.assertEqual(dataset["__excel_context__"]["hotel_name"], "测试酒店")
        self.assertEqual(dataset["__excel_context__"]["period_start"], "2026-06-16")
        self.assertEqual(dataset["__excel_context__"]["period_end"], "2026-07-15")
        self.assertEqual(len(dataset["hotel_daily"]), 1)
        self.assertEqual(dataset["hotel_daily"][0]["room_revenue"], 4060)
        self.assertEqual(dataset["ota_funnel"][0]["platform"], "meituan")
        self.assertEqual(dataset["ota_funnel"][0]["hos_score_rank"], 2)
        self.assertEqual(dataset["hotel_performance_daily"][0]["value_month"], 90.5)

    def test_legacy_workbook_behavior_is_preserved(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "hotel_daily"
        sheet.append(["business_date", "room_count", "room_nights"])
        sheet.append([date(2026, 7, 15), 32, 28])

        dataset = load_excel_dataset(self._save(workbook))

        self.assertNotIn("__excel_context__", dataset)
        self.assertEqual(len(dataset["hotel_daily"]), 1)
        self.assertEqual(dataset["hotel_daily"][0]["room_count"], 32)

    def test_excel_only_fields_are_displayed_without_inventing_scores(self):
        result = {
            "visual_diagnosis": {
                "items": [
                    {"standard_item_id": 7, "data_status": "missing", "item_score": None, "fields": []},
                    {"standard_item_id": 22, "data_status": "manual_pending", "item_score": None, "fields": []},
                ]
            }
        }
        raw = {
            "scan_orders": [
                {"order_id": "A", "order_count": 1},
                {"order_id": "B", "order_count": 2},
            ],
            "manual_inputs": [
                {"crown_type": "普通挂冠", "operator": "张三", "recorded_at": "2026-07-15 18:00:00"}
            ],
        }

        enriched = enrich_customer_excel_result(result, raw)
        items = {item["standard_item_id"]: item for item in enriched["visual_diagnosis"]["items"]}

        self.assertEqual(items[7]["fields"][0]["value"], 3)
        self.assertIsNone(items[7]["item_score"])
        self.assertEqual(items[22]["fields"][0]["value"], "普通挂冠")
        self.assertIsNone(items[22]["item_score"])


if __name__ == "__main__":
    unittest.main()
