from __future__ import annotations

from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


TEMPLATE_FILE_NAME = "s14_ota_diagnosis_template.xlsx"

DARK = "155B4D"
GREEN = "1E8F67"
HEADER = "2B6F63"
LIGHT_GREEN = "E6F5EF"
SAMPLE = "F2FAF7"
INPUT = "FFF5CC"
ORANGE = "FFF0D8"
WHITE = "FFFFFF"
TEXT = "26343D"
MUTED = "667784"
BORDER = "D7E4DF"

YES_NO = ["是", "否"]
STATUS = ["OPEN", "CLOSED", "PENDING"]


def _previous_year(day: date) -> date:
    try:
        return day.replace(year=day.year - 1)
    except ValueError:
        return day.replace(year=day.year - 1, day=28)


def _block(
    sheet: str,
    title: str,
    module: str,
    page: str,
    headers: list[str],
    samples: list[list[Any]],
    blank_rows: int,
    note: str,
    validations: dict[str, list[Any]] | None = None,
) -> dict[str, Any]:
    return {
        "sheet": sheet,
        "title": title,
        "module": module,
        "page": page,
        "headers": headers,
        "samples": samples,
        "blank_rows": blank_rows,
        "note": note,
        "validations": validations or {},
    }


def _blocks() -> list[dict[str, Any]]:
    today = date.today()
    start = today - timedelta(days=29)
    previous = _previous_year(today)
    stamp = datetime.combine(today, datetime.min.time()).replace(hour=18)
    demo_id = "demo_hotel"
    demo_name = "示例酒店（请替换）"

    flow_headers = [
        "是否导入", "系统模块（勿改）", "酒店ID", "业务日期", "平台", "统计周期",
        "曝光人数", "同行曝光", "曝光排名", "浏览人数", "同行浏览", "浏览排名",
        "支付订单", "同行订单", "订单排名", "曝光-浏览转化率", "同行曝光-浏览转化率",
        "转化率排名", "浏览-支付转化率", "同行浏览-支付转化率", "支付转化排名",
        "HOS分", "HOS排名", "信息分", "快照时间",
    ]

    return [
        _block(
            "02_基础经营", "A. 酒店基本信息", "basic_info", "报告基本信息",
            ["是否导入", "系统模块（勿改）", "酒店ID", "酒店名称", "诊断开始日期", "诊断结束日期"],
            [["否", "basic_info", demo_id, demo_name, start, today]], 2,
            "只保留一行真实酒店信息；酒店ID在所有区域中保持一致。",
        ),
        _block(
            "02_基础经营", "B. 日度经营数据", "hotel_daily", "报告顶部经营概况",
            ["是否导入", "系统模块（勿改）", "酒店ID", "酒店名称", "业务日期", "总房量", "售出间夜", "房费收入", "平均房价", "出租率", "快照时间"],
            [["否", "hotel_daily", demo_id, demo_name, today, 32, 28, 4060, 145, .875, stamp]], 31,
            "建议填写最近30天；未知值留空，不要用0代替。",
        ),
        _block(
            "02_基础经营", "C. 经营同比数据", "hotel_performance_daily", "页面01 月度经营趋势YOY",
            ["是否导入", "系统模块（勿改）", "酒店ID", "酒店名称", "业务日期", "指标名称", "当日值", "本月值", "本年值", "快照时间"],
            [
                ["否", "hotel_performance_daily", demo_id, demo_name, today, "房费", 554, 5540, 62540, stamp],
                ["否", "hotel_performance_daily", demo_id, demo_name, today, "出租率", .92, .905, .863, stamp],
                ["否", "hotel_performance_daily", demo_id, demo_name, today, "平均房价", 148, 145, 142.5, stamp],
                ["否", "hotel_performance_daily", demo_id, demo_name, today, "RevPAR", 136.2, 131.3, 123, stamp],
                ["否", "hotel_performance_daily", demo_id, demo_name, previous, "房费", 810, 9240, 70120, datetime.combine(previous, datetime.min.time()).replace(hour=18)],
                ["否", "hotel_performance_daily", demo_id, demo_name, previous, "出租率", .88, .872, .846, datetime.combine(previous, datetime.min.time()).replace(hour=18)],
                ["否", "hotel_performance_daily", demo_id, demo_name, previous, "平均房价", 152, 149.5, 146.2, datetime.combine(previous, datetime.min.time()).replace(hour=18)],
                ["否", "hotel_performance_daily", demo_id, demo_name, previous, "RevPAR", 133.8, 130.4, 123.7, datetime.combine(previous, datetime.min.time()).replace(hour=18)],
            ], 16,
            "固定4个指标；本期和去年同期各4行。",
            {"指标名称": ["房费", "出租率", "平均房价", "RevPAR"]},
        ),
        _block(
            "03_房型商品", "A. 房型经营表现", "room_type_performance_daily", "页面02 房型RevPAR与低效房型",
            ["是否导入", "系统模块（勿改）", "酒店ID", "业务日期", "房型ID", "房型名称", "指标名称", "当日值", "近30天值", "本年值", "快照时间"],
            [
                ["否", "room_type_performance_daily", demo_id, today, "R01", "至臻·电竞大床房", "出租率", .92, .885, .84, stamp],
                ["否", "room_type_performance_daily", demo_id, today, "R01", "至臻·电竞大床房", "RevPAR", 155, 137.2, 118, stamp],
                ["否", "room_type_performance_daily", demo_id, today, "R02", "轻奢·电竞大床房", "出租率", .55, .48, .70, stamp],
                ["否", "room_type_performance_daily", demo_id, today, "R02", "轻奢·电竞大床房", "RevPAR", 138, 66.2, 98, stamp],
            ], 20,
            "每个房型至少两行：出租率和RevPAR。",
            {"指标名称": ["出租率", "RevPAR"]},
        ),
        _block(
            "03_房型商品", "B. 平台商品与房型", "products", "页面11 房型名称卖点优化",
            ["是否导入", "系统模块（勿改）", "酒店ID", "快照时间", "平台", "商品ID", "商品名称", "房型ID", "房型名称", "挂牌价", "售卖价"],
            [
                ["否", "products", demo_id, stamp, "美团", "P1001", "电竞大床房·连住优惠", "R01", "至臻·电竞大床房", 168, 158],
                ["否", "products", demo_id, stamp, "美团", "P1002", "电竞双床房·双人开黑", "R02", "至臻·电竞双床房", 188, 178],
            ], 30,
            "一行一个在售商品。",
            {"平台": ["美团", "携程", "飞猪", "去哪儿", "抖音"]},
        ),
        _block(
            "04_流量用户", "A. 曝光来源", "exposure_daily", "页面03 曝光数据分析",
            ["是否导入", "系统模块（勿改）", "酒店ID", "业务日期", "整体曝光", "非广告曝光", "广告曝光", "广告曝光占比", "快照时间"],
            [["否", "exposure_daily", demo_id, today, 2200, 1750, 450, 450 / 2200, stamp]], 31,
            "建议填写最近30天；整体曝光应等于非广告曝光加广告曝光。",
        ),
        _block(
            "04_流量用户", "B. 流量、HOS和信息分", "ota_funnel", "页面04/06/08 流量、HOS、信息分",
            flow_headers,
            [["否", "ota_funnel", demo_id, today, "美团", "日", 2200, 2250, "3/20", 225, 205, "4/20", 18, 20, "5/20", .1023, .0911, "6/20", .08, .0976, "7/20", 5.35, "2/20", 95, stamp]], 31,
            "最关键模块；至少填写最近7天。",
            {"平台": ["美团"], "统计周期": ["日"]},
        ),
        _block(
            "04_流量用户", "C. 用户来源", "user_source_monthly", "页面05 用户来源数据",
            ["是否导入", "系统模块（勿改）", "酒店ID", "统计月份", "本地占比", "异地占比", "新客占比", "老客占比", "快照时间"],
            [["否", "user_source_monthly", demo_id, today.strftime("%Y-%m"), .62, .38, .71, .29, stamp]], 12,
            "每月一行；两组占比各自应约等于100%。",
        ),
        _block(
            "05_推广竞对", "A. 推广支出", "promotion_finance", "页面09 推广数据",
            ["是否导入", "系统模块（勿改）", "酒店ID", "交易时间", "交易类型", "交易金额", "快照时间"],
            [["否", "promotion_finance", demo_id, f"{today} 00:00:00-23:59:59", "推广通支出", -100, stamp]], 30,
            "一笔交易一行；金额只填数字。",
            {"交易类型": ["推广通支出"]},
        ),
        _block(
            "05_推广竞对", "B. 推广对应收入", "promotion_revenue", "页面09 推广ROI",
            ["是否导入", "系统模块（勿改）", "酒店ID", "统计月份", "维度类型", "维度名称", "房费收入", "快照时间"],
            [["否", "promotion_revenue", demo_id, today.strftime("%Y-%m"), "渠道", "美团EBK", 53358.04, stamp]], 12,
            "用于计算推广ROI。",
            {"维度类型": ["渠道"], "维度名称": ["美团EBK"]},
        ),
        _block(
            "05_推广竞对", "C. 竞对订单流失", "order_loss_monthly", "页面12 竞争对手分析",
            ["是否导入", "系统模块（勿改）", "酒店ID", "统计月份", "竞对酒店名称", "流失订单数", "流失金额", "流失房型", "关注状态", "快照时间"],
            [["否", "order_loss_monthly", demo_id, today.strftime("%Y-%m"), "示例竞对酒店A", 8, 1460, "电竞大床房、双床房", 1, stamp]], 20,
            "关注状态：1=已关注，0=未关注。",
            {"关注状态": [0, 1]},
        ),
        _block(
            "06_口碑权益", "A. 点评概览", "review_overviews", "页面13 口碑分析",
            ["是否导入", "系统模块（勿改）", "酒店ID", "点评平台", "点评评分", "点评总数", "未回复点评数", "差评数", "快照时间"],
            [
                ["否", "review_overviews", demo_id, "美团", 4.70, 3268, 12, 49, stamp],
                ["否", "review_overviews", demo_id, "大众点评", 4.65, 1250, 6, 22, stamp],
            ], 6,
            "每个平台一行，无需逐条填写评论。",
            {"点评平台": ["美团", "大众点评", "携程", "飞猪", "去哪儿"]},
        ),
        _block(
            "06_口碑权益", "B. 权益中心与页面基础", "joined_rights", "页面10/14 页面基础与权益中心",
            ["是否导入", "系统模块（勿改）", "酒店ID", "酒店名称", "权益名称", "有效房型范围", "热门商圈词命中", "快照时间"],
            [["否", "joined_rights", demo_id, demo_name, "会员专享价", "全部在售房型", "花溪公园", stamp]], 15,
            "一项权益一行。",
        ),
        _block(
            "07_配置视频挂冠", "A. 权益与配置状态", "promotion_status", "页面15-21及23 配置状态",
            ["是否导入", "系统模块（勿改）", "酒店ID", "配置项名称", "开通状态", "报名状态", "生效状态", "快照时间"],
            [["否", "promotion_status", demo_id, name, "PENDING", "PENDING", "PENDING", stamp] for name in ["优美乐", "商旅专享价", "公益流量", "钟点房", "酒店亮点", "预约开票", "自动接单"]], 0,
            "固定7行；OPEN=已开通，CLOSED=未开通，PENDING=不确定。",
            {"配置项名称": ["优美乐", "商旅专享价", "公益流量", "钟点房", "酒店亮点", "预约开票", "自动接单"], "开通状态": STATUS, "报名状态": STATUS, "生效状态": STATUS},
        ),
        _block(
            "07_配置视频挂冠", "B. 视频上传状态", "video_upload_status", "页面21 首页视频",
            ["是否导入", "系统模块（勿改）", "酒店ID", "视频类型", "已上传数量", "需上传数量", "快照时间"],
            [
                ["否", "video_upload_status", demo_id, "room_type_video", 3, 4, stamp],
                ["否", "video_upload_status", demo_id, "hotel_preview_video", 1, 1, stamp],
                ["否", "video_upload_status", demo_id, "room_type_preview_video", 2, 3, stamp],
            ], 0,
            "固定3种视频类型；英文值不要修改。",
            {"视频类型": ["room_type_video", "hotel_preview_video", "room_type_preview_video"]},
        ),
        _block(
            "07_配置视频挂冠", "C. 酒店挂冠", "manual_inputs", "页面22 酒店挂冠",
            ["是否导入", "系统模块（勿改）", "酒店ID", "挂冠类型", "录入人", "录入时间"],
            [["否", "manual_inputs", demo_id, "普通挂冠", "张三", stamp]], 2,
            "只填写1行；评分阈值未冻结时仅展示不擅自计分。",
            {"挂冠类型": ["黑金挂冠", "普通挂冠", "无挂冠"]},
        ),
        _block(
            "08_可选补充", "A. 扫码订单", "scan_orders", "页面07 扫码订单数据",
            ["是否导入", "系统模块（勿改）", "酒店ID", "扫码时间", "订单ID", "扫码订单数量"],
            [["否", "scan_orders", demo_id, stamp, "SCAN-DEMO-001", 1]], 30,
            "可选；订单ID不要重复。评分阈值未冻结时仅展示。",
        ),
        _block(
            "08_可选补充", "B. 周边活动", "nearby_events", "经营建议补充",
            ["是否导入", "系统模块（勿改）", "酒店ID", "活动名称", "开始日期", "结束日期", "活动地址", "距离公里", "活动类型", "预计需求"],
            [["否", "nearby_events", demo_id, "示例电竞赛事", today + timedelta(days=5), today + timedelta(days=6), "示例体育馆", 3.2, "电竞赛事", "预计周末需求上升"]], 15,
            "可选；用于远期价格和套餐建议。",
        ),
        _block(
            "08_可选补充", "C. 竞品价格", "competitors", "竞品价格分析补充",
            ["是否导入", "系统模块（勿改）", "酒店ID", "业务日期", "平台", "竞品名称", "房型名称", "价格", "排名", "评分", "已售房量", "活动标签"],
            [["否", "competitors", demo_id, today, "美团", "示例竞对酒店A", "电竞大床房", 158, 5, 4.72, 12, "会员价"]], 30,
            "可选；一行一个竞品房型。",
            {"平台": ["美团", "携程", "飞猪", "去哪儿"]},
        ),
    ]


def _style_sheet(sheet, subtitle: str) -> None:
    sheet.merge_cells("A1:L1")
    sheet["A1"] = sheet.title
    sheet["A1"].fill = PatternFill("solid", fgColor=DARK)
    sheet["A1"].font = Font(color=WHITE, bold=True, size=15)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 32
    sheet.merge_cells("A2:L2")
    sheet["A2"] = subtitle
    sheet["A2"].fill = PatternFill("solid", fgColor=LIGHT_GREEN)
    sheet["A2"].font = Font(color=MUTED, size=10)
    sheet["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[2].height = 36
    sheet.freeze_panes = "A3"


def _apply_validation(sheet, column: int, start: int, end: int, values: list[Any]) -> None:
    quoted = ",".join(str(value) for value in values)
    validation = DataValidation(type="list", formula1=f'"{quoted}"', allow_blank=True)
    sheet.add_data_validation(validation)
    validation.add(f"{get_column_letter(column)}{start}:{get_column_letter(column)}{end}")


def _write_block(sheet, start_row: int, block: dict[str, Any]) -> tuple[int, str]:
    headers = block["headers"]
    last_column = len(headers)
    last_letter = get_column_letter(last_column)
    border = Border(*(Side(style="thin", color=BORDER) for _ in range(4)))

    sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=last_column)
    title_cell = sheet.cell(start_row, 1, f'{block["title"]} 〔系统模块：{block["module"]}〕')
    title_cell.fill = PatternFill("solid", fgColor=GREEN)
    title_cell.font = Font(color=WHITE, bold=True, size=11)
    title_cell.alignment = Alignment(vertical="center")

    sheet.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=last_column)
    note_cell = sheet.cell(start_row + 1, 1, block["note"])
    note_cell.fill = PatternFill("solid", fgColor=LIGHT_GREEN)
    note_cell.font = Font(color=MUTED, size=10)
    note_cell.alignment = Alignment(wrap_text=True, vertical="center")

    header_row = start_row + 2
    for column, value in enumerate(headers, start=1):
        cell = sheet.cell(header_row, column, value)
        cell.fill = PatternFill("solid", fgColor=HEADER)
        cell.font = Font(color=WHITE, bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    sheet.row_dimensions[header_row].height = 30

    sample_start = header_row + 1
    for row_offset, values in enumerate(block["samples"]):
        row = sample_start + row_offset
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row, column, value)
            cell.fill = PatternFill("solid", fgColor=SAMPLE)
            cell.font = Font(color=TEXT, size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border

    blank_start = sample_start + len(block["samples"])
    blank_end = blank_start + block["blank_rows"] - 1
    for row in range(blank_start, blank_end + 1):
        sheet.cell(row, 1, "否")
        for column in range(1, last_column + 1):
            cell = sheet.cell(row, column)
            cell.fill = PatternFill("solid", fgColor=INPUT)
            cell.font = Font(color=TEXT, size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border

    data_end = max(sample_start + len(block["samples"]) - 1, blank_end)
    _apply_validation(sheet, 1, sample_start, data_end, YES_NO)
    sheet.conditional_formatting.add(
        f"A{sample_start}:A{data_end}",
        FormulaRule(formula=[f'A{sample_start}="是"'], fill=PatternFill("solid", fgColor="DFF4EA")),
    )

    header_index = {name: index + 1 for index, name in enumerate(headers)}
    for header, values in block["validations"].items():
        if header in header_index:
            _apply_validation(sheet, header_index[header], sample_start, data_end, values)

    for index, header in enumerate(headers, start=1):
        width = 15
        if header == "系统模块（勿改）":
            width = 22
        elif any(token in header for token in ("名称", "房型", "内容", "说明", "范围", "地址")):
            width = 25
        elif any(token in header for token in ("日期", "时间")):
            width = 19
        sheet.column_dimensions[get_column_letter(index)].width = width

    return data_end + 2, f"'{sheet.title}'!A{sample_start}:A{data_end}"


def build_excel_template(path: str | Path) -> Path:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)

    guide = workbook.create_sheet("00_填写说明")
    guide.merge_cells("A1:F1")
    guide["A1"] = "S14 酒店诊断最佳客户填报模板"
    guide["A1"].fill = PatternFill("solid", fgColor=DARK)
    guide["A1"].font = Font(color=WHITE, bold=True, size=15)
    guide["A1"].alignment = Alignment(vertical="center")
    guide.row_dimensions[1].height = 34
    guide.merge_cells("A2:F2")
    guide["A2"] = "绿色行为示例，黄色行为新增区域。示例行默认不导入；覆盖真实数据后，将是否导入改为“是”。工作表名称、系统模块列和表头不要修改。"
    guide["A2"].fill = PatternFill("solid", fgColor=ORANGE)
    guide["A2"].font = Font(color="8A4B00", bold=True, size=10)
    guide["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    guide.row_dimensions[2].height = 42
    rules = [
        ["序号", "填写规则", "正确做法", "错误示例", "是否必须", "影响"],
        [1, "示例数据", "覆盖demo_hotel和示例酒店", "保留示例值直接上传", "必须", "会生成示例报告"],
        [2, "是否导入", "真实行改为是，空行保持否", "全部空行改为是", "必须", "产生空数据"],
        [3, "比例", "填0.6129或61.29%", "填61.29但不带%", "必须", "比例放大100倍"],
        [4, "经营同比", "本期与去年同期都填写", "只填本期", "必须", "YOY无法计算"],
        [5, "未知数据", "保持空白", "用0表示未知", "否", "可能误判经营差"],
    ]
    for row in rules:
        guide.append(row)
    for cell in guide[3]:
        cell.fill = PatternFill("solid", fgColor=HEADER)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in guide.iter_rows(min_row=4, max_row=guide.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.border = Border(*(Side(style="thin", color=BORDER) for _ in range(4)))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column, width in enumerate([8, 24, 38, 30, 12, 28], start=1):
        guide.column_dimensions[get_column_letter(column)].width = width
    guide.freeze_panes = "A3"

    blocks = _blocks()
    sheet_subtitles = {
        "02_基础经营": "酒店基础信息、日度经营、本期与去年同期经营数据。",
        "03_房型商品": "房型出租率、RevPAR，以及平台在售商品和价格。",
        "04_流量用户": "曝光、浏览、支付订单、同行均值/排名、HOS分、信息分和用户来源。",
        "05_推广竞对": "推广投入、渠道收入、推广ROI和订单流失竞对。",
        "06_口碑权益": "点评平台概览、权益中心和页面基础信息。",
        "07_配置视频挂冠": "固定配置状态、视频数量和酒店挂冠。",
        "08_可选补充": "扫码订单、周边活动和竞品价格；没有数据可以全部留空。",
    }
    progress: list[tuple[dict[str, Any], str]] = []
    for sheet_name, subtitle in sheet_subtitles.items():
        sheet = workbook.create_sheet(sheet_name)
        _style_sheet(sheet, subtitle)
        row = 4
        for block in [item for item in blocks if item["sheet"] == sheet_name]:
            row, import_range = _write_block(sheet, row, block)
            progress.append((block, import_range))

    dashboard = workbook.create_sheet("01_填报总览", 1)
    dashboard.merge_cells("A1:G1")
    dashboard["A1"] = "S14 客户填报进度与页面覆盖"
    dashboard["A1"].fill = PatternFill("solid", fgColor=DARK)
    dashboard["A1"].font = Font(color=WHITE, bold=True, size=15)
    dashboard.merge_cells("A2:G2")
    dashboard["A2"] = "真实数据行改为“是”后，下表自动统计已导入行数。未填写的模块不会阻止报告生成，但对应页面会显示数据缺口。"
    dashboard["A2"].fill = PatternFill("solid", fgColor=LIGHT_GREEN)
    dashboard["A2"].font = Font(color=MUTED, size=10)
    dashboard["A2"].alignment = Alignment(wrap_text=True)
    dashboard.append([])
    dashboard.append(["序号", "页面/模块", "填写位置", "最低建议行数", "已导入行数", "填报状态", "说明"])
    for cell in dashboard[4]:
        cell.fill = PatternFill("solid", fgColor=HEADER)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    required_by_module = {
        "basic_info": 1, "hotel_daily": 7, "hotel_performance_daily": 8,
        "room_type_performance_daily": 2, "products": 1, "exposure_daily": 7,
        "ota_funnel": 7, "user_source_monthly": 1, "promotion_finance": 1,
        "promotion_revenue": 1, "order_loss_monthly": 1, "review_overviews": 1,
        "joined_rights": 1, "promotion_status": 7, "video_upload_status": 3,
        "manual_inputs": 1, "scan_orders": 0, "nearby_events": 0, "competitors": 0,
    }
    for index, (block, import_range) in enumerate(progress, start=1):
        row = dashboard.max_row + 1
        required = required_by_module.get(block["module"], 1)
        dashboard.append([index, block["title"], block["sheet"], required, None, None, block["page"]])
        dashboard.cell(row, 5, f'=COUNTIF({import_range},"是")')
        if required == 0:
            dashboard.cell(row, 6, f'=IF(E{row}>0,"已补充","可选")')
        else:
            dashboard.cell(row, 6, f'=IF(E{row}>={required},"已填写",IF(E{row}>0,"部分填写","未填写"))')
        for cell in dashboard[row]:
            cell.border = Border(*(Side(style="thin", color=BORDER) for _ in range(4)))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column, width in enumerate([8, 30, 22, 14, 14, 16, 34], start=1):
        dashboard.column_dimensions[get_column_letter(column)].width = width
    dashboard.freeze_panes = "A4"

    workbook.save(output_path)
    return output_path


def ensure_excel_template(output_root: str | Path) -> Path:
    template_path = Path(output_root) / "_templates" / TEMPLATE_FILE_NAME
    build_excel_template(template_path)
    return template_path
