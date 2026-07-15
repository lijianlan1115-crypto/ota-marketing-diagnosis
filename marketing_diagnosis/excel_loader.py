from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from marketing_diagnosis.customer_excel_loader_v2 import (
    is_customer_excel_template,
    load_customer_excel_workbook,
)
from marketing_diagnosis.data import section_for_sheet


def _rows(sheet) -> list[dict[str, Any]]:
    raw_rows = list(sheet.iter_rows(values_only=True))
    if not raw_rows:
        return []
    headers = [str(cell).strip() if cell is not None else "" for cell in raw_rows[0]]
    result: list[dict[str, Any]] = []
    for raw in raw_rows[1:]:
        item = {headers[i]: value for i, value in enumerate(raw) if i < len(headers) and headers[i]}
        if any(value not in (None, "") for value in item.values()):
            result.append(item)
    return result


def _load_legacy_workbook(workbook) -> dict[str, list[dict[str, Any]]]:
    """Keep the original one-sheet-per-section Excel behavior unchanged."""
    dataset: dict[str, list[dict[str, Any]]] = {}
    for sheet_name in workbook.sheetnames:
        section = section_for_sheet(sheet_name)
        if not section:
            continue
        dataset.setdefault(section, []).extend(_rows(workbook[sheet_name]))
    return dataset


def load_excel_dataset(path: str | Path) -> dict[str, list[dict[str, Any]]]:
    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Excel file not found: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        # New customer template: multiple Chinese data blocks can live in one
        # worksheet. Detection is based on its fixed control headers and does
        # not affect the original English/legacy workbook layout.
        if is_customer_excel_template(workbook):
            return load_customer_excel_workbook(workbook)
        return _load_legacy_workbook(workbook)
    finally:
        workbook.close()
