from __future__ import annotations

import re
import unicodedata
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SECTIONS = (
    "hotel_performance_daily",
    "room_type_performance_daily",
    "exposure_daily",
    "ota_funnel",
    "user_source_monthly",
    "promotion_finance",
    "promotion_revenue",
    "products",
    "order_loss_monthly",
    "review_overviews",
    "joined_rights",
    "promotion_status",
    "video_upload_status",
    "manual_inputs",
)


def _text(value: Any) -> str:
    return unicodedata.normalize("NFKC", str(value or "")).strip()


def _key(value: Any) -> str:
    return re.sub(r"[\s_\-—（）()·:：/]+", "", _text(value)).lower()


def _sheet_key(value: Any) -> str:
    text = re.sub(r"^\s*\d+\s*[_\-.、 ]*", "", _text(value))
    return _key(text)


SHEET_KIND = {
    _sheet_key("00_基本信息"): "meta",
    _sheet_key("01_月度经营YOY"): "hotel_performance_daily",
    _sheet_key("02_房型表现"): "room_type_performance_daily",
    _sheet_key("03_曝光来源"): "exposure_daily",
    _sheet_key("04_流量指标"): "flow_long",
    _sheet_key("05_用户来源"): "user_source_monthly",
    _sheet_key("06_扫码订单"): "scan_orders",
    _sheet_key("07_推广财务"): "promotion_finance",
    _sheet_key("08_渠道月收入"): "promotion_revenue",
    _sheet_key("09_商品房型"): "products",
    _sheet_key("10_竞对流失"): "order_loss_monthly",
    _sheet_key("11_点评概览"): "review_overviews",
    _sheet_key("12_权益中心"): "joined_rights",
    _sheet_key("13_配置状态"): "promotion_status",
    _sheet_key("14_视频状态"): "video_upload_status",
    _sheet_key("15_人工录入"): "manual_inputs",
}


COMMON = {
    "是否导入": "__include__",
    "酒店id": "hotel_id",
    "酒店编号": "hotel_id",
}


RAW_HEADER_MAPS: dict[str, dict[str, str]] = {
    "meta": {
        "是否导入": "__include__",
        "酒店id": "hotel_id",
        "酒店编号": "hotel_id",
        "酒店名称": "hotel_name",
        "诊断渠道": "platform",
        "平台": "platform",
        "诊断开始日期": "period_start",
        "开始日期": "period_start",
        "诊断结束日期": "period_end",
        "结束日期": "period_end",
    },
    "hotel_performance_daily": {
        **COMMON,
        "业务日期": "business_date",
        "指标名称": "metric_name",
        "当日值": "value_day",
        "本月值": "value_month",
        "本年值": "value_year",
        "快照时间": "snapshot_time",
    },
    "room_type_performance_daily": {
        **COMMON,
        "业务日期": "business_date",
        "房型id": "room_type_id",
        "房型名称": "room_type_name",
        "pms房型id": "pms_rate_room_type_id",
        "已售间夜": "room_nights",
        "出租率": "occupancy_rate",
        "房型收入": "room_revenue",
        "平均房价": "adr",
        "revpar": "revpar",
        "快照时间": "snapshot_time",
    },
    "exposure_daily": {
        **COMMON,
        "业务日期": "business_date",
        "整体曝光": "total_exposure",
        "非广告曝光": "non_ad_exposure",
        "广告曝光": "ad_exposure",
        "广告曝光占比": "ad_exposure_ratio_pct",
        "快照时间": "snapshot_time",
    },
    "flow_long": {
        **COMMON,
        "业务日期": "business_date",
        "平台": "platform",
        "统计周期类型": "period_type",
        "指标代码": "metric_code",
        "指标名称": "metric_name",
        "指标值": "metric_value",
        "同行均值": "peer_average",
        "同行排名": "competitor_rank",
        "快照时间": "snapshot_time",
    },
    "user_source_monthly": {
        **COMMON,
        "统计月份": "period_month",
        "本地占比": "local_user_pct",
        "异地占比": "nonlocal_user_pct",
        "新客占比": "new_user_pct",
        "老客占比": "returning_user_pct",
        "快照时间": "snapshot_time",
    },
    "scan_orders": {
        **COMMON,
        "扫码时间": "scan_time",
        "订单id": "order_id",
        "扫码订单数量": "scan_order_count",
    },
    "promotion_finance": {
        **COMMON,
        "交易时间": "transaction_time",
        "交易类型": "transaction_type",
        "交易金额": "transaction_amount",
        "快照时间": "snapshot_time",
    },
    "promotion_revenue": {
        **COMMON,
        "统计月份": "period_month",
        "维度类型": "dimension_type",
        "维度名称": "dimension_name",
        "房费收入": "room_revenue",
        "快照时间": "snapshot_time",
    },
    "products": {
        **COMMON,
        "业务日期": "business_date",
        "平台": "platform",
        "房型id": "room_type_id",
        "房型名称": "room_type_name",
        "商品id": "ota_product_id",
        "商品名称": "product_name",
        "挂牌价": "listed_price",
        "售卖价": "final_price",
        "快照时间": "snapshot_time",
    },
    "order_loss_monthly": {
        **COMMON,
        "统计月份": "period_month",
        "竞对酒店名称": "competitor_hotel_name",
        "流失订单数": "competitor_loss_order_count",
        "流失金额": "competitor_loss_amount",
        "流失房型": "lost_room_types_text",
        "关注状态": "follow_status",
        "快照时间": "snapshot_time",
    },
    "review_overviews": {
        **COMMON,
        "点评平台": "review_platform",
        "点评评分": "rating_avg",
        "点评总数": "review_count",
        "未回复点评数": "unreplied_review_count",
        "快照时间": "snapshot_time",
    },
    "joined_rights": {
        **COMMON,
        "酒店名称": "hotel_name",
        "权益名称": "right_name",
        "有效房型范围": "effective_room_scope",
        "快照时间": "snapshot_time",
    },
    "promotion_status": {
        **COMMON,
        "配置项名称": "promotion_name",
        "开通状态": "status",
        "报名状态": "enroll_status",
        "生效状态": "effective_status",
        "快照时间": "snapshot_time",
    },
    "video_upload_status": {
        **COMMON,
        "视频类型": "video_type",
        "已上传数量": "uploaded_count",
        "需上传数量": "required_count",
        "快照时间": "snapshot_time",
    },
    "manual_inputs": {
        **COMMON,
        "挂冠类型": "crown_type",
        "录入人": "operator",
        "录入时间": "recorded_at",
    },
}

HEADER_MAPS = {
    kind: {_key(header): canonical for header, canonical in mapping.items()}
    for kind, mapping in RAW_HEADER_MAPS.items()
}

NUMERIC_FIELDS = {
    "value_day", "value_month", "value_year", "room_nights", "occupancy_rate",
    "room_revenue", "adr", "revpar", "total_exposure", "non_ad_exposure",
    "ad_exposure", "ad_exposure_ratio_pct", "metric_value", "peer_average",
    "local_user_pct", "nonlocal_user_pct", "new_user_pct", "returning_user_pct",
    "scan_order_count", "transaction_amount", "listed_price", "final_price",
    "competitor_loss_order_count", "competitor_loss_amount", "rating_avg",
    "review_count", "unreplied_review_count", "uploaded_count", "required_count",
}

DATE_FIELDS = {"business_date", "period_start", "period_end"}
DATETIME_FIELDS = {"snapshot_time", "transaction_time", "scan_time", "recorded_at"}

PLATFORM_MAP = {
    "美团": "meituan", "大众点评": "dianping", "携程": "ctrip",
    "飞猪": "fliggy", "去哪儿": "qunar", "同程": "tongcheng",
    "多渠道": "multi", "全部渠道": "multi",
}

REVIEW_PLATFORM_MAP = {
    "美团": "meituan", "美团酒店": "meituan", "大众点评": "dianping",
    "点评": "dianping",
}

FLOW_METRICS = {
    _key("曝光人数"): ("exposure", "peer_exposure", "exposure_rank"),
    _key("浏览人数"): ("views", "peer_views", "views_rank"),
    _key("支付订单数"): ("paid_orders", "peer_paid_orders", "paid_orders_rank"),
    _key("曝光-浏览转化率"): (
        "exposure_to_view_rate", "peer_exposure_to_view_rate", "exposure_to_view_rate_rank"
    ),
    _key("浏览-支付转化率"): (
        "payment_conversion_rate", "peer_payment_conversion_rate", "payment_conversion_rate_rank"
    ),
    _key("支付转化率"): (
        "payment_conversion_rate", "peer_payment_conversion_rate", "payment_conversion_rate_rank"
    ),
    _key("HOS分"): ("hos_score", None, "hos_score_rank"),
    _key("信息分"): ("content_score", None, None),
}


def _number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    text = _text(value).replace(",", "").replace("￥", "").replace("¥", "")
    percent = text.endswith("%")
    if percent:
        text = text[:-1]
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    number = float(match.group())
    return number / 100 if percent else number


def _date_text(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return _text(value)[:10]


def _datetime_text(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    return _text(value)


def _platform(value: Any) -> str:
    text = _text(value)
    return PLATFORM_MAP.get(text, text.lower())


def _included(value: Any) -> bool:
    text = _key(value)
    return text not in {"否", "no", "false", "0", "示例", "不导入"}


def _find_header_row(sheet, kind: str) -> tuple[int, list[str]] | None:
    aliases = HEADER_MAPS[kind]
    best: tuple[int, int, list[str]] | None = None
    for row_index, raw in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        headers = [_text(cell) for cell in raw]
        score = sum(1 for header in headers if _key(header) in aliases)
        if score and (best is None or score > best[0]):
            best = (score, row_index, headers)
    if best is None:
        return None
    return best[1], best[2]


def _sheet_rows(sheet, kind: str) -> list[dict[str, Any]]:
    found = _find_header_row(sheet, kind)
    if found is None:
        return []
    header_row, headers = found
    aliases = HEADER_MAPS[kind]
    rows: list[dict[str, Any]] = []
    for raw in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        mapped: dict[str, Any] = {}
        for index, value in enumerate(raw):
            if index >= len(headers):
                break
            canonical = aliases.get(_key(headers[index]))
            if canonical:
                mapped[canonical] = value
        if not mapped or not any(value not in (None, "") for value in mapped.values()):
            continue
        if not _included(mapped.pop("__include__", "是")):
            continue
        rows.append(_convert_row(kind, mapped))
    return rows


def _convert_row(kind: str, row: dict[str, Any]) -> dict[str, Any]:
    converted: dict[str, Any] = {}
    for key, value in row.items():
        if key in NUMERIC_FIELDS:
            converted[key] = _number(value)
        elif key in DATE_FIELDS:
            converted[key] = _date_text(value)
        elif key in DATETIME_FIELDS:
            converted[key] = _datetime_text(value)
        elif key == "platform":
            converted[key] = _platform(value)
        elif key == "review_platform":
            text = _text(value)
            converted[key] = REVIEW_PLATFORM_MAP.get(text, text.lower())
        elif key in {"status", "enroll_status", "effective_status"}:
            converted[key] = _text(value).upper()
        else:
            converted[key] = _text(value) if isinstance(value, str) else value
    return converted


def _in_period(value: Any, start: str | None, end: str | None) -> bool:
    day = _date_text(value)
    if not day:
        return not start and not end
    return (not start or day >= start) and (not end or day <= end)


def _flow_rows(rows: list[dict[str, Any]], default_platform: str) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, str], dict[str, Any]] = {}
    stamps: dict[tuple[tuple[str, str, str], str], str] = {}
    for row in rows:
        definition = FLOW_METRICS.get(_key(row.get("metric_name")))
        if not definition:
            continue
        business_date = _date_text(row.get("business_date"))
        platform = _platform(row.get("platform")) or default_platform or "meituan"
        if platform == "multi":
            platform = "meituan"
        period_type = _text(row.get("period_type")) or "日"
        group_key = (business_date, platform, period_type)
        target = grouped.setdefault(
            group_key,
            {
                "hotel_id": row.get("hotel_id"),
                "business_date": business_date,
                "platform": platform,
                "period_type": period_type,
                "snapshot_time": _datetime_text(row.get("snapshot_time")),
                "source_table": "Excel：04_流量指标",
            },
        )
        value_key, peer_key, rank_key = definition
        stamp = _datetime_text(row.get("snapshot_time"))
        metric_stamp_key = (group_key, value_key)
        if stamp >= stamps.get(metric_stamp_key, ""):
            target[value_key] = _number(row.get("metric_value"))
            if peer_key:
                target[peer_key] = _number(row.get("peer_average"))
            if rank_key and row.get("competitor_rank") not in (None, ""):
                rank = _text(row.get("competitor_rank"))
                target[rank_key] = rank
                target[f"{rank_key}_raw"] = rank
            stamps[metric_stamp_key] = stamp
        if stamp > str(target.get("snapshot_time") or ""):
            target["snapshot_time"] = stamp
    return list(grouped.values())


def _scan_summary(rows: list[dict[str, Any]], start: str | None, end: str | None) -> dict[str, Any] | None:
    total = 0.0
    matched = 0
    for row in rows:
        scan_time = row.get("scan_time")
        if scan_time not in (None, "") and not _in_period(scan_time, start, end):
            continue
        explicit = _number(row.get("scan_order_count"))
        if explicit is not None:
            total += explicit
            matched += 1
        elif row.get("order_id") not in (None, "") or scan_time not in (None, ""):
            total += 1
            matched += 1
    if matched == 0:
        return None
    return {
        "platform": "meituan",
        "period_type": "scan_order_summary",
        "scan_order_count": total,
        "scan_order_date_column": "scan_time",
        "scan_order_period_start": start,
        "scan_order_period_end": end,
        "source_table": "Excel：06_扫码订单",
    }


def _promotion_summary(rows: list[dict[str, Any]], start: str | None, end: str | None) -> list[dict[str, Any]]:
    values: list[float] = []
    for row in rows:
        if not _in_period(row.get("transaction_time"), start, end):
            continue
        transaction_type = "".join(char for char in _text(row.get("transaction_type")) if not char.isspace())
        if transaction_type != "推广通支出":
            continue
        value = _number(row.get("transaction_amount"))
        if value is not None:
            values.append(abs(value))
    if not values:
        return []
    return [{
        "transaction_type": "推广通支出",
        "transaction_amount": sum(values),
        "promotion_spend_summary": True,
        "transaction_count": len(values),
        "period_start": start,
        "period_end": end,
        "source_table": "Excel：07_推广财务",
    }]


def _meta(rows: list[dict[str, Any]]) -> dict[str, Any]:
    if not rows:
        return {}
    row = rows[0]
    return {
        "hotel_id": _text(row.get("hotel_id")),
        "hotel_name": _text(row.get("hotel_name")),
        "platform": _platform(row.get("platform")) or "multi",
        "period_start": _date_text(row.get("period_start")),
        "period_end": _date_text(row.get("period_end")),
    }


def load_excel_package(
    path: str | Path,
    *,
    period_start: str | None = None,
    period_end: str | None = None,
) -> tuple[dict[str, list[dict[str, Any]]], dict[str, Any]]:
    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Excel file not found: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    raw_by_kind: dict[str, list[dict[str, Any]]] = defaultdict(list)
    diagnostics: dict[str, Any] = {"mode": "excel_chinese_v2", "sheets": {}}

    for sheet_name in workbook.sheetnames:
        kind = SHEET_KIND.get(_sheet_key(sheet_name))
        if not kind:
            continue
        rows = _sheet_rows(workbook[sheet_name], kind)
        raw_by_kind[kind].extend(rows)
        diagnostics["sheets"][sheet_name] = {"kind": kind, "rows": len(rows), "status": "ok"}

    meta = _meta(raw_by_kind.get("meta") or [])
    start = period_start or meta.get("period_start") or None
    end = period_end or meta.get("period_end") or None
    default_platform = str(meta.get("platform") or "multi")

    dataset: dict[str, list[dict[str, Any]]] = {section: [] for section in SECTIONS}
    for section in SECTIONS:
        if section in {"ota_funnel", "promotion_finance"}:
            continue
        dataset[section] = list(raw_by_kind.get(section) or [])
        for row in dataset[section]:
            row.setdefault("source_table", f"Excel：{section}")

    dataset["ota_funnel"] = _flow_rows(raw_by_kind.get("flow_long") or [], default_platform)
    scan = _scan_summary(raw_by_kind.get("scan_orders") or [], start, end)
    if scan:
        dataset["ota_funnel"].append(scan)
    dataset["promotion_finance"] = _promotion_summary(raw_by_kind.get("promotion_finance") or [], start, end)

    hotel_id = meta.get("hotel_id")
    for rows in dataset.values():
        for row in rows:
            if hotel_id and row.get("hotel_id") in (None, ""):
                row["hotel_id"] = hotel_id

    diagnostics["period_start"] = start
    diagnostics["period_end"] = end
    diagnostics["sections"] = {section: len(rows) for section, rows in dataset.items()}
    dataset["__source_diagnostics__"] = [{
        "source": str(workbook_path),
        "mode": "excel_chinese_v2",
        "tables": {
            sheet_name: {"status": info["status"], "rows": info["rows"]}
            for sheet_name, info in diagnostics["sheets"].items()
        },
        "transformations": [
            {"section": "ota_funnel", "rule": "中文长表指标转为页面宽表"},
            {"section": "promotion_finance", "rule": "交易类型Unicode标准化后按周期汇总推广通支出绝对值"},
            {"section": "manual_inputs", "rule": "挂冠类型进入第23项评分"},
        ],
    }]
    return dataset, meta


def load_excel_dataset(path: str | Path) -> dict[str, list[dict[str, Any]]]:
    dataset, _ = load_excel_package(path)
    return dataset


__all__ = ["load_excel_dataset", "load_excel_package"]
